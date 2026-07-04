#!/usr/bin/env python3
"""
Phase 5 — Generate the 9 ØØT Excel templates from `templates/excel/SPEC.md`.

Produces:
- X1 partner-output-ledger.xlsx
- X2 reward-species-declaration.xlsx
- X3 business-review.xlsx
- X4 klarna-test.xlsx
- X5 metr-baseline.xlsx
- X6 agent-skill-roi.xlsx
- X7 eu-ai-act-mapping.xlsx
- X8 treasury-runway.xlsx (optional)
- X9 oot-readiness.xlsx

Each workbook:
- Sheets in spec order.
- Named ranges per spec.
- Formulas as specified verbatim.
- Conditional formatting per spec.
- Sample data (3-5 rows, [SAMPLE] prefix).
- Hidden `_metadata` sheet (spec_version, generation_date, oot_version, file_id).
- README sheet last.

Run from repo root: `python3 scripts/build_excel.py`.
Output: `templates/excel/*.xlsx`.
"""
from __future__ import annotations

from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule, FormulaRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.worksheet import Worksheet

# ──────────────────────────────────────────────────────────────────────────────
# Constants
# ──────────────────────────────────────────────────────────────────────────────

SPEC_VERSION = "1.0.0"
OOT_VERSION = "1.0.0"
GENERATION_DATE = date.today().isoformat()
TEMPLATE_DIR = Path(__file__).parent.parent / "templates" / "excel"

# Colours used across files
RED_FILL = PatternFill("solid", fgColor="FFC7CE")
YELLOW_FILL = PatternFill("solid", fgColor="FFEB9C")
GREEN_FILL = PatternFill("solid", fgColor="C6EFCE")
HEADER_FILL = PatternFill("solid", fgColor="305496")
HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _add_metadata_sheet(wb: Workbook, file_id: str) -> None:
    """Add a hidden _metadata sheet with file provenance."""
    ws = wb.create_sheet("_metadata")
    ws.sheet_state = "hidden"
    ws["A1"] = "key"
    ws["B1"] = "value"
    ws["A2"] = "spec_version"
    ws["B2"] = SPEC_VERSION
    ws["A3"] = "generation_date"
    ws["B3"] = GENERATION_DATE
    ws["A4"] = "oot_version"
    ws["B4"] = OOT_VERSION
    ws["A5"] = "file_id"
    ws["B5"] = file_id


def _style_header_row(ws: Worksheet, row: int = 1, last_col: int = 26) -> None:
    """Apply standard header styling to a row."""
    for col in range(1, last_col + 1):
        cell = ws.cell(row=row, column=col)
        if cell.value is None:
            continue
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGN


def _add_readme_sheet(wb: Workbook, content: str) -> None:
    """Add a README sheet at the end with the given content."""
    ws = wb.create_sheet("README")
    for i, line in enumerate(content.splitlines(), start=1):
        ws.cell(row=i, column=1, value=line)
    ws.column_dimensions["A"].width = 110


def _set_column_widths(ws: Worksheet, widths: dict[str, int]) -> None:
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width


# ──────────────────────────────────────────────────────────────────────────────
# X1 — partner-output-ledger.xlsx
# ──────────────────────────────────────────────────────────────────────────────

def build_x1() -> None:
    """X1 partner-output-ledger.xlsx — daily output capture, monthly variable."""
    wb = Workbook()
    wb.remove(wb.active)  # remove default sheet

    # ── Sheet 1: Output_Log ──
    ws = wb.create_sheet("Output_Log")
    headers = [
        "log_id", "date", "partner_id", "output_type", "output_ref",
        "output_spec_ref", "value_tier", "ai_authored_pct", "rework_within_30d",
        "partner_multiplier", "value_envelope", "computed_variable", "notes",
    ]
    for col_idx, h in enumerate(headers, start=1):
        ws.cell(row=1, column=col_idx, value=h)

    # Value envelope lookup table in O1:P5 (per SPEC)
    ws["O1"] = "value_tier"
    ws["P1"] = "base_envelope"
    envelope_table = [("XS", 100), ("L", 500), ("M", 2000), ("S", 8000)]
    for i, (tier, env) in enumerate(envelope_table, start=2):
        ws.cell(row=i, column=15, value=tier)
        ws.cell(row=i, column=16, value=env)

    # Sample data (3 rows)
    sample_data = [
        ("OL-20260301-001", "2026-03-01", "[SAMPLE]mira-tek", "pr_merged", "gh:1100",
         "[[partners/mira-tek/output-specs/2026-02-25--login-flow]]", "M", 30, "No",
         1.0, None, None, "Initial sample data"),
        ("OL-20260302-001", "2026-03-02", "[SAMPLE]davor-krznar", "review_completed", "gh:1108",
         "[[partners/davor-krznar/output-specs/2026-02-28--code-review-cohort]]", "L", 0, "No",
         1.0, None, None, ""),
        ("OL-20260303-001", "2026-03-03", "[SAMPLE]anya-gorska", "deal_closed", "Acme Corp Q1",
         "[[partners/anya-gorska/output-specs/2026-02-15--acme-deal]]", "S", 0, "No",
         1.0, None, None, "First closed deal"),
    ]
    for row_idx, row in enumerate(sample_data, start=2):
        for col_idx, val in enumerate(row, start=1):
            ws.cell(row=row_idx, column=col_idx, value=val)

    # Formulas (rows 2-4 sample data)
    for row_idx in range(2, 5):
        # K (value_envelope) = VLOOKUP(G, O1:P5, 2, FALSE)
        ws.cell(row=row_idx, column=11,
                value=f"=VLOOKUP(G{row_idx},$O$2:$P$5,2,FALSE)")
        # L (computed_variable) = K * J * IF(I="Yes", 0, 1)
        ws.cell(row=row_idx, column=12,
                value=f'=K{row_idx}*J{row_idx}*IF(I{row_idx}="Yes",0,1)')

    # Named ranges
    wb.defined_names["output_log"] = DefinedName(
        name="output_log", attr_text="Output_Log!$A:$M")
    wb.defined_names["value_envelope_table"] = DefinedName(
        name="value_envelope_table", attr_text="Output_Log!$O$1:$P$5")

    # Conditional formatting: column I red if "Yes"
    ws.conditional_formatting.add(
        "I2:I1000",
        CellIsRule(operator="equal", formula=['"Yes"'], fill=RED_FILL))
    # Column L green gradient
    ws.conditional_formatting.add(
        "L2:L1000",
        ColorScaleRule(start_type="min", start_color="FFFFFF",
                       end_type="max", end_color="63BE7B"))

    _style_header_row(ws, last_col=16)
    _set_column_widths(ws, {
        "A": 22, "B": 12, "C": 18, "D": 18, "E": 24, "F": 40,
        "G": 10, "H": 14, "I": 16, "J": 12, "K": 12, "L": 16, "M": 30,
        "O": 12, "P": 14,
    })

    # ── Sheet 2: Monthly_Variable ──
    ws2 = wb.create_sheet("Monthly_Variable")
    mv_headers = [
        "month", "partner_id", "total_outputs", "total_variable", "base_pay",
        "total_compensation", "sign_off_status", "approval_date", "payment_date",
    ]
    for col_idx, h in enumerate(mv_headers, start=1):
        ws2.cell(row=1, column=col_idx, value=h)

    # Sample row
    ws2.append([
        "2026-03", "[SAMPLE]mira-tek", 4, 3100, 1371, None, "draft", None, None
    ])
    ws2["F2"] = "=D2+E2"

    # Sign-off status validation
    dv = DataValidation(
        type="list",
        formula1='"draft,partner_reviewed,partner_disputed,partner_unresponsive,founder_approved,paid"',
        allow_blank=True)
    ws2.add_data_validation(dv)
    dv.add("G2:G1000")

    _style_header_row(ws2, last_col=9)
    _set_column_widths(ws2, {
        "A": 10, "B": 18, "C": 14, "D": 14, "E": 12, "F": 18,
        "G": 22, "H": 14, "I": 14,
    })

    # ── Sheet 3: Partner_Dashboard ──
    ws3 = wb.create_sheet("Partner_Dashboard")
    pd_headers = [
        "partner_id", "rolling_30d_outputs", "rolling_30d_variable",
        "ytd_variable", "current_month_forecast",
    ]
    for col_idx, h in enumerate(pd_headers, start=1):
        ws3.cell(row=1, column=col_idx, value=h)
    ws3.append(["[SAMPLE]mira-tek", 4, 3100, 3100, 4500])
    _style_header_row(ws3, last_col=5)
    _set_column_widths(ws3, {"A": 18, "B": 20, "C": 20, "D": 14, "E": 24})

    # ── Sheet 4: Partner_Multipliers_Snapshot ──
    ws4 = wb.create_sheet("Partner_Multipliers_Snapshot")
    pms_headers = ["partner_id", "full_name", "base_amount", "output_multiplier",
                   "reward_species", "snapshot_date"]
    for col_idx, h in enumerate(pms_headers, start=1):
        ws4.cell(row=1, column=col_idx, value=h)
    ws4.append(["[SAMPLE]mira-tek", "Mira Tek", 30000, 1.0, "hybrid", "2026-03-01"])
    ws4.append(["[SAMPLE]davor-krznar", "Davor Krznar", 30000, 1.0, "eat-what-you-kill", "2026-03-01"])
    ws4.append(["[SAMPLE]anya-gorska", "Anya Gorska", 24000, 1.0, "hybrid", "2026-03-01"])
    # Conditional formatting on snapshot_date — red if older than 35 days
    ws4.conditional_formatting.add(
        "F2:F1000",
        FormulaRule(formula=[f"AND(F2<>\"\",TODAY()-F2>35)"], fill=RED_FILL))
    _style_header_row(ws4, last_col=6)
    _set_column_widths(ws4, {"A": 18, "B": 22, "C": 14, "D": 18, "E": 20, "F": 14})

    # ── Sheet 5: README ──
    readme = """# partner-output-ledger.xlsx

PURPOSE
Records daily output per partner. Source of truth for monthly variable pay (Routine R3) and weekly Business Review agenda (Routine R2).

WRITTEN BY
- Routine R1 (Daily Output Capture, daily 18:00) — appends rows to Output_Log.

READ BY
- Routine R2 (Weekly BR Prep) — surfaces notable outputs, blockers, KPI movements.
- Routine R3 (Monthly Variable Calc) — locks previous month's Output_Log; populates Monthly_Variable.
- Routine R4 (Quarterly Long-Tail Settlement) — references Output_Log for outputs eligible for long-tail entitlements.

REVIEWED BY
- Each partner: monthly, before Variable sign-off.
- Founder: monthly, for Variable approval.
- Friday Business Review: weekly, in the standing agenda.

DESIGN DECISION — column J (partner_multiplier) is NOT a formula
The X1.J value is written by R1 at row-append time. R1 reads X2's output_multiplier at runtime and writes the resolved number. This avoids the unreliable cross-workbook formula pattern in openpyxl. The Partner_Multipliers_Snapshot sheet is a read-only mirror refreshed monthly by R3 — it is not consumed by any formula.

DESIGN DECISION — AI-authored output is paid at full rate at month-1
The L column formula does NOT discount by ai_authored_pct. The framework's correction discipline is the rework-within-30d zero-out. AI-assisted output requiring rework "was not actually output, it was just typing."

DO NOT
- Edit Output_Log rows that have been included in a paid Monthly_Variable. Use the renegotiation flow per Skill Pack S3.
- Bypass the rework_within_30d field. The YOLO model's correction discipline depends on it.
- Aggregate values in ways that obscure individual partner attribution.
- Modify column J via formula — the cross-workbook discipline requires Routine writes.
"""
    _add_metadata_sheet(wb, "X1")
    _add_readme_sheet(wb, readme)

    out_path = TEMPLATE_DIR / "partner-output-ledger.xlsx"
    wb.save(out_path)
    print(f"  ✓ {out_path.name}")


# ──────────────────────────────────────────────────────────────────────────────
# X2 — reward-species-declaration.xlsx
# ──────────────────────────────────────────────────────────────────────────────

def build_x2() -> None:
    """X2 reward-species-declaration.xlsx — per-partner contract."""
    wb = Workbook()
    wb.remove(wb.active)

    # Sample partner sheet
    partner_id = "[SAMPLE]mira-tek"

    # ── Sheet 1: Partner_Profile ──
    ws = wb.create_sheet("Partner_Profile")
    pp_headers = [
        "partner_id", "full_name", "cohort", "start_date", "jurisdiction",
        "base_currency", "stablecoin_upgrade_pref", "unit_fund_interest",
        "two_worlds_self_id",
    ]
    for col_idx, h in enumerate(pp_headers, start=1):
        ws.cell(row=1, column=col_idx, value=h)
    ws.append([
        partner_id, "Mira Tek", "full-time-partner", "2026-03-15", "SI",
        "EUR", "Yes", "Yes", "agentic-engineer"
    ])
    # Validation
    dv_cohort = DataValidation(
        type="list",
        formula1='"full-time-partner,project-specialist,advisor"',
        allow_blank=False)
    ws.add_data_validation(dv_cohort)
    dv_cohort.add("C2:C1000")
    dv_two = DataValidation(
        type="list",
        formula1='"vibe-coder,agentic-engineer,transitional,non-code"',
        allow_blank=False)
    ws.add_data_validation(dv_two)
    dv_two.add("I2:I1000")

    _style_header_row(ws, last_col=9)
    _set_column_widths(ws, {
        "A": 20, "B": 22, "C": 22, "D": 12, "E": 12,
        "F": 14, "G": 24, "H": 18, "I": 22,
    })

    # ── Sheet 2: Base_Variable_Split ──
    ws2 = wb.create_sheet("Base_Variable_Split")
    bvs_headers = [
        "reward_species", "base_amount", "variable_weight_personal",
        "variable_weight_team", "variable_weight_company", "output_multiplier",
        "bonus_split_personal", "bonus_split_team", "bonus_split_company",
    ]
    for col_idx, h in enumerate(bvs_headers, start=1):
        ws2.cell(row=1, column=col_idx, value=h)
    ws2.append(["hybrid", 30000, 0.6, 0.3, 0.1, 1.0, 0.34, 0.33, 0.33])

    # Validation: variable weights C+D+E = 1.0; bonus splits G+H+I = 1.0
    # Conditional formatting: highlight red if sum != 1.0
    # Guard with COUNT>0 so blank rows (sum 0) are not painted red — only populated rows.
    ws2.conditional_formatting.add(
        "C2:E1000",
        FormulaRule(formula=["AND(COUNT($C2:$E2)>0,ROUND(SUM($C2:$E2),3)<>1)"], fill=RED_FILL))
    ws2.conditional_formatting.add(
        "G2:I1000",
        FormulaRule(formula=["AND(COUNT($G2:$I2)>0,ROUND(SUM($G2:$I2),3)<>1)"], fill=RED_FILL))

    dv_species = DataValidation(
        type="list",
        formula1='"eat-what-you-kill,lockstep,hybrid"',
        allow_blank=False)
    ws2.add_data_validation(dv_species)
    dv_species.add("A2:A1000")

    _style_header_row(ws2, last_col=9)
    _set_column_widths(ws2, dict.fromkeys("ABCDEFGHI", 20))

    # ── Sheet 3: Long_Tail_Schedule ──
    ws3 = wb.create_sheet("Long_Tail_Schedule")
    lt_headers = [
        "output_id", "description", "partner_share_pct", "settlement_period",
        "start_date", "end_date", "total_settled_to_date",
    ]
    for col_idx, h in enumerate(lt_headers, start=1):
        ws3.cell(row=1, column=col_idx, value=h)
    ws3.append([
        "[SAMPLE]OL-20260212-014", "Acme contract — 3yr ARR", 8, "quarterly",
        "2026-04-01", "2029-03-31", 0,
    ])
    _style_header_row(ws3, last_col=7)
    _set_column_widths(ws3, {
        "A": 24, "B": 32, "C": 18, "D": 18, "E": 12, "F": 12, "G": 22,
    })

    # ── Sheet 4: Unit_Fund_Eligibility (locked in Gen 1) ──
    ws4 = wb.create_sheet("Unit_Fund_Eligibility")
    uf_headers = ["credit_balance", "credits_earned_to_date",
                  "credits_consumed", "units_held", "last_subscription_date"]
    for col_idx, h in enumerate(uf_headers, start=1):
        ws4.cell(row=1, column=col_idx, value=h)
    ws4.append([0, 0, 0, 0, None])
    # Add header note
    ws4["A4"] = "This sheet activates in Generation 2."
    ws4["A5"] = "Fields locked / read-only in Gen 1. Activation requires written re-affirmation per the Charter."
    ws4.protection.sheet = True
    ws4.protection.password = "oot-gen1-locked"  # password is documented; users intentionally need to bypass to test Gen 2 flows
    _style_header_row(ws4, last_col=5)

    # ── Sheet 5: Renegotiation_Log ──
    ws5 = wb.create_sheet("Renegotiation_Log")
    rl_headers = [
        "renegotiation_date", "initiated_by", "reason", "fields_changed",
        "approved_by", "brain_link",
    ]
    for col_idx, h in enumerate(rl_headers, start=1):
        ws5.cell(row=1, column=col_idx, value=h)
    _style_header_row(ws5, last_col=6)
    _set_column_widths(ws5, {"A": 18, "B": 16, "C": 30, "D": 30, "E": 18, "F": 32})

    # ── README ──
    readme = """# reward-species-declaration.xlsx

PURPOSE
Per-partner compensation contract. Records reward species, base amount, variable weights, output multiplier, bonus splits, long-tail entitlement schedule, and renegotiation history.

For organisations <20 partners: ONE workbook with one sheet per partner. For larger orgs: one workbook per partner.

UPDATED BY
- Manual at onboarding (Skill Pack S3 §4.1).
- Manual at renegotiation (Skill Pack S3 §4.1 renegotiation flow).
- Routine R4 (Quarterly Long-Tail Settlement) updates Long_Tail_Schedule total_settled_to_date.

REVIEWED BY
- Quarterly partner check-in (Skill Pack S5 §4.6).

VALIDATION
- Variable weights (Base_Variable_Split columns C+D+E) MUST sum to 1.0.
- Bonus splits (G+H+I) MUST sum to 1.0.
- Conditional formatting flags violations in red.
- Column F (output_multiplier) is NOT part of the variable-weight sum; it is a free-running multiplier.

DO NOT
- Override fields without a signed renegotiation flow per Skill Pack S3 §4.1.
- Bypass the jurisdiction field; legal touchpoints depend on it (see governance/EU-AI-ACT.md and docs/06-when-to-call-a-lawyer.md).
- Alter Long_Tail_Schedule retroactively without a new Renegotiation_Log entry.
- Activate Unit_Fund_Eligibility before Gen 2 (sheet is password-protected by design).
"""
    _add_metadata_sheet(wb, "X2")
    _add_readme_sheet(wb, readme)

    out_path = TEMPLATE_DIR / "reward-species-declaration.xlsx"
    wb.save(out_path)
    print(f"  ✓ {out_path.name}")


# ──────────────────────────────────────────────────────────────────────────────
# X3 — business-review.xlsx
# ──────────────────────────────────────────────────────────────────────────────

def build_x3() -> None:
    """X3 business-review.xlsx — Friday BR working document."""
    wb = Workbook()
    wb.remove(wb.active)

    # ── Sheet 1: Weekly_Review ──
    ws = wb.create_sheet("Weekly_Review")
    wr_headers = [
        "week_starting", "notable_outputs", "blockers", "klarna_test_status",
        "kpi_movements", "decisions_due", "meeting_notes", "brain_link",
    ]
    for col_idx, h in enumerate(wr_headers, start=1):
        ws.cell(row=1, column=col_idx, value=h)
    ws.append([
        "2026-04-13",
        "[SAMPLE] Davor S-tier onboarding rebuild merged; Anya signed Acme MSA; Mira payment-rail PR.",
        "[SAMPLE] Anya: Acme legal review (ETA 2026-04-22). Mira: Stripe deprecation.",
        "[SAMPLE] KT-2026-014 proceeded (16/20). KT-2026-015 in scoring.",
        "[SAMPLE] Customer count 84→87. Runway 18.2→17.9 months. ai_skill_roi 4.2→4.6.",
        "[SAMPLE] D-2026-04-007 Unit Fund pilot Q3; D-2026-04-009 Stripe→Adyen.",
        "(filled during meeting)",
        "[[business-reviews/2026-04-17]]",
    ])
    _style_header_row(ws, last_col=8)
    _set_column_widths(ws, {
        "A": 14, "B": 60, "C": 50, "D": 40, "E": 40, "F": 40, "G": 40, "H": 30,
    })

    # ── Sheet 2: Monthly_BR ──
    ws2 = wb.create_sheet("Monthly_BR")
    mb_headers = [
        "month", "total_outputs", "total_variable_paid", "gross_margin",
        "treasury_runway_months", "customer_count_delta", "partner_count",
        "klarna_tests_proceeded", "klarna_tests_held", "ai_skill_roi",
    ]
    for col_idx, h in enumerate(mb_headers, start=1):
        ws2.cell(row=1, column=col_idx, value=h)
    ws2.append(["[SAMPLE]2026-03", 47, 28500, 0.62, 18.2, 3, 8, 1, 0, 4.2])
    _style_header_row(ws2, last_col=10)

    # ── Sheet 3: Decisions_Log ──
    ws3 = wb.create_sheet("Decisions_Log")
    dl_headers = [
        "decision_id", "date", "decision", "accountable", "consulted",
        "brain_link", "reversal_threshold", "review_date",
    ]
    for col_idx, h in enumerate(dl_headers, start=1):
        ws3.cell(row=1, column=col_idx, value=h)
    ws3.append([
        "D-2026-04-007", "2026-04-17",
        "[SAMPLE] Defer Unit Fund pilot until counsel sign-off.",
        "founder-tali", "davor,anya,mira",
        "[[decisions/D-2026-04-007]]",
        "If counsel signs off before 2026-06-15, revisit at next BR.",
        "2026-06-15",
    ])
    _style_header_row(ws3, last_col=8)
    _set_column_widths(ws3, {"A": 16, "B": 12, "C": 50, "D": 18, "E": 30, "F": 30, "G": 50, "H": 14})

    # ── Sheet 4: Blockers ──
    ws4 = wb.create_sheet("Blockers")
    bl_headers = [
        "blocker_id", "raised_date", "description", "partner_blocked",
        "owner", "status", "resolved_date", "brain_link",
    ]
    for col_idx, h in enumerate(bl_headers, start=1):
        ws4.cell(row=1, column=col_idx, value=h)
    ws4.append([
        "[SAMPLE]B-2026-04-001", "2026-04-15",
        "Acme legal review of long-tail clause", "anya-gorska", "anya-gorska",
        "open", None, "[[blockers/2026-04-15-acme-legal]]",
    ])
    dv_status = DataValidation(type="list", formula1='"open,resolved"', allow_blank=False)
    ws4.add_data_validation(dv_status)
    dv_status.add("F2:F1000")
    _style_header_row(ws4, last_col=8)

    # ── Sheet 5: Klarna_Test_Hits ──
    ws5 = wb.create_sheet("Klarna_Test_Hits")
    kt_headers = ["test_id", "summary", "status", "score", "decision", "review_date_90d"]
    for col_idx, h in enumerate(kt_headers, start=1):
        ws5.cell(row=1, column=col_idx, value=h)
    ws5.append([
        "[SAMPLE]KT-2026-014", "Replace manual customer-onboarding email with AI drafter",
        "proceeded", 16, "PROCEED", "2026-08-12",
    ])
    _style_header_row(ws5, last_col=6)

    # ── README ──
    readme = """# business-review.xlsx

PURPOSE
Friday Business Review working document. Generated by Routine R2; reviewed in the BR; outcomes committed to the Brain.

WRITTEN BY
- Routine R2 (Weekly BR Prep, Friday 08:00) — populates Weekly_Review row.
- Scribe during/after the meeting — fills meeting_notes; appends decisions_log entries; updates blockers.

READ BY
- BR participants pre-meeting (Friday 09:00 onwards).
- Skill Pack S5 (Reporting & Business Review) at all times.

DO NOT
- Skip the Klarna status block even if no open tests exist.
- Auto-publish BR summaries externally — they may contain sensitive partner-level data.
- Take a decision in the BR without creating a corresponding firm/decisions/D-YYYY-NNN.md page during the meeting.
"""
    _add_metadata_sheet(wb, "X3")
    _add_readme_sheet(wb, readme)

    out_path = TEMPLATE_DIR / "business-review.xlsx"
    wb.save(out_path)
    print(f"  ✓ {out_path.name}")


# ──────────────────────────────────────────────────────────────────────────────
# X4 — klarna-test.xlsx
# ──────────────────────────────────────────────────────────────────────────────

def build_x4() -> None:
    """X4 klarna-test.xlsx — Klarna Test scoring."""
    wb = Workbook()
    wb.remove(wb.active)

    # ── Sheet 1: Decision_Log ──
    ws = wb.create_sheet("Decision_Log")
    dl_headers = [
        "test_id", "date", "decision_summary", "trigger", "trigger_ref",
        "scorer", "non_beneficiary_reviewer", "total_score", "decision",
        "reversal_plan_ref", "review_date_90d", "post_review_outcome",
    ]
    for col_idx, h in enumerate(dl_headers, start=1):
        ws.cell(row=1, column=col_idx, value=h)
    ws.append([
        "[SAMPLE]KT-2026-014", "2026-04-29",
        "Replace manual customer-onboarding email step with AI-drafted version",
        "pr_label", "https://github.com/firm/firm-saas/pull/1502",
        "davor-krznar", "mira-tek",
        None, None, None, date(2026, 8, 12), None,
    ])
    # H total_score: VLOOKUP from Klarna_Score
    ws["H2"] = "=IFERROR(VLOOKUP(A2,Klarna_Score!A:L,12,FALSE),\"\")"
    # I decision: =IF(H>=14,"PROCEED","HOLD")
    ws["I2"] = '=IF(ISBLANK(H2),"",IF(H2>=14,"PROCEED","HOLD"))'
    # K review_date_90d (R7 contract: column K = the 90-day review date). Written as a REAL
    # date cell (not text) so the review-overdue CF's `K2<TODAY()` comparison actually fires —
    # a text date sorts greater than any number and would silently never trigger.
    ws["K2"].number_format = "yyyy-mm-dd"

    # Conditional formatting
    ws.conditional_formatting.add(
        "I2:I1000",
        CellIsRule(operator="equal", formula=['"PROCEED"'], fill=GREEN_FILL))
    ws.conditional_formatting.add(
        "I2:I1000",
        CellIsRule(operator="equal", formula=['"HOLD"'], fill=RED_FILL))
    # Review-overdue formatting on K: past the review date AND no post_review_outcome (L) yet.
    # Guarded on K non-blank so empty rows aren't flagged.
    ws.conditional_formatting.add(
        "K2:K1000",
        FormulaRule(formula=["AND(K2<>\"\",K2<TODAY(),L2=\"\")"], fill=RED_FILL))

    dv_trigger = DataValidation(type="list",
                                formula1='"pr_label,manual,pre_rollout"', allow_blank=False)
    ws.add_data_validation(dv_trigger)
    dv_trigger.add("D2:D1000")

    _style_header_row(ws, last_col=12)
    _set_column_widths(ws, {
        "A": 16, "B": 12, "C": 50, "D": 14, "E": 40, "F": 16,
        "G": 22, "H": 12, "I": 12, "J": 28, "K": 14, "L": 30,
    })

    # ── Sheet 2: Klarna_Score ──
    ws2 = wb.create_sheet("Klarna_Score")
    qs_headers = ["test_id"] + [f"Q{i}" for i in range(1, 11)] + [
        "total", "evidence_links", "scorer_signoff", "non_beneficiary_signoff",
    ]
    for col_idx, h in enumerate(qs_headers, start=1):
        ws2.cell(row=1, column=col_idx, value=h)
    ws2.append([
        "[SAMPLE]KT-2026-014",
        2, 2, 2, 2, 2, 2, 2, 2, 2, 2,  # all 2s
        None,  # total (formula)
        "All evidence in Brain page firm/klarna-tests/KT-2026-014",
        "Yes", "Yes",
    ])
    # L (total) = SUM(B:K)
    ws2["L2"] = "=SUM(B2:K2)"

    # Validation: each Q column accepts only 0, 1, 2 (no n/a — per B4 correction)
    for col_letter in "BCDEFGHIJK":
        dv_q = DataValidation(type="list", formula1='"0,1,2"', allow_blank=False,
                              prompt="Score: 0 (not addressed), 1 (partial), 2 (fully addressed). See governance/KLARNA-TEST.md.",
                              promptTitle="Klarna Question Score")
        ws2.add_data_validation(dv_q)
        dv_q.add(f"{col_letter}2:{col_letter}1000")

    # Conditional formatting on Q columns: red 0 / yellow 1 / green 2
    for col_letter in "BCDEFGHIJK":
        rng = f"{col_letter}2:{col_letter}1000"
        ws2.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=["0"], fill=RED_FILL))
        ws2.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=["1"], fill=YELLOW_FILL))
        ws2.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=["2"], fill=GREEN_FILL))
    # L (total) red if <14, green if >=14 — guarded on populated row (test_id A non-blank)
    # so blank rows (SUM formula absent → treated as 0) are not painted red.
    ws2.conditional_formatting.add(
        "L2:L1000",
        FormulaRule(formula=["AND($A2<>\"\",L2<14)"], fill=RED_FILL))
    ws2.conditional_formatting.add(
        "L2:L1000",
        FormulaRule(formula=["AND($A2<>\"\",L2>=14)"], fill=GREEN_FILL))

    dv_yn = DataValidation(type="list", formula1='"Yes,No"', allow_blank=False)
    ws2.add_data_validation(dv_yn)
    dv_yn.add("N2:O1000")

    _style_header_row(ws2, last_col=15)

    # ── Sheet 3: Quality_Gates ──
    ws3 = wb.create_sheet("Quality_Gates")
    qg_headers = ["test_id", "metric", "baseline", "threshold", "current", "breach_check"]
    for col_idx, h in enumerate(qg_headers, start=1):
        ws3.cell(row=1, column=col_idx, value=h)
    ws3.append([
        "[SAMPLE]KT-2026-014", "first-contact resolution %",
        82, 75, 79, None,
    ])
    ws3["F2"] = '=IF(E2<D2,"BREACHED","OK")'
    ws3.conditional_formatting.add(
        "F2:F1000",
        CellIsRule(operator="equal", formula=['"BREACHED"'], fill=RED_FILL))
    _style_header_row(ws3, last_col=6)

    # ── Sheet 4: Reversal_Plan ──
    ws4 = wb.create_sheet("Reversal_Plan")
    rp_headers = [
        "test_id", "reversal_action", "standby_partners",
        "standby_contract_ref", "reversal_owner", "max_reversal_time",
    ]
    for col_idx, h in enumerate(rp_headers, start=1):
        ws4.cell(row=1, column=col_idx, value=h)
    ws4.append([
        "[SAMPLE]KT-2026-014",
        "Restore manual customer-onboarding email path; high-tier customers escalate to Anya",
        "anya-gorska",
        "(retained as full-time-partner; no separate standby contract needed)",
        "founder-tali", "2 weeks",
    ])
    _style_header_row(ws4, last_col=6)

    # ── README ──
    readme = """# klarna-test.xlsx

PURPOSE
Operational scoring of the framework's signature epistemic check. Triggered by Routine R7 on `ai-replaces-human` PR labels; also used manually for non-PR decisions.

WRITTEN BY
- Routine R7 (Klarna Test Trigger) — appends rows to Decision_Log on PR-label events.
- Manual scoring by the assigned scorer + non-beneficiary reviewer in Klarna_Score.

READ BY
- Routine R7 — the `oot/klarna-test` GitHub status check workflow polls Klarna_Score for the matching test_id; sets the check passing only when total_score >= 14 AND scorer_signoff = Yes AND non_beneficiary_signoff = Yes.
- Routine R6 (Daily Audit Trail) — references Klarna Test entries.
- Routine R2 (Weekly BR Prep) — surfaces open + recently-resolved tests in the Klarna_Test_Hits mirror in X3.

VALIDATION
- Each Q column accepts only 0, 1, or 2.
- No "n/a" affordance — every question always applies. Q8 (public-communication posture) explicitly requires "no comms" to be itself a written, owned posture to score 2.

DO NOT
- Bypass the discipline. The Klarna Test is non-negotiable.
- Score with a beneficiary as the non-beneficiary reviewer (Q7).
- Flip total_score >= 14 by rationalising 1s into 2s. The pre-committed remediation flow exists for tests that score below threshold.
- Modify completed (signed-off) Klarna_Score rows. Re-scores after remediation update the SAME row; do not create a new test_id.
"""
    _add_metadata_sheet(wb, "X4")
    _add_readme_sheet(wb, readme)

    out_path = TEMPLATE_DIR / "klarna-test.xlsx"
    wb.save(out_path)
    print(f"  ✓ {out_path.name}")


# ──────────────────────────────────────────────────────────────────────────────
# X5 — metr-baseline.xlsx
# ──────────────────────────────────────────────────────────────────────────────

def build_x5() -> None:
    """X5 metr-baseline.xlsx — pre-rollout productivity baseline."""
    wb = Workbook()
    wb.remove(wb.active)

    # ── Sheet 1: Baseline_Metrics ──
    ws = wb.create_sheet("Baseline_Metrics")
    bm_headers = [
        "metric_name", "source", "baseline_value", "capture_date",
        "capture_period_days", "owner", "notes",
    ]
    for col_idx, h in enumerate(bm_headers, start=1):
        ws.cell(row=1, column=col_idx, value=h)

    canonical_metrics = [
        ("[DORA] deployment_frequency", "GitHub API", 2.4, "2026-02-28", 90, "tomislav-novak", "per-engineer weekly average"),
        ("[DORA] lead_time_for_changes", "GitHub PR data", 3.7, "2026-02-28", 90, "tomislav-novak", "PR open → merge in days (median)"),
        ("[DORA] change_failure_rate", "Sentry + revert tags", 0.11, "2026-02-28", 90, "tomislav-novak", "fraction"),
        ("[DORA] time_to_restore_service", "Incident tracker", 4.1, "2026-02-28", 90, "tomislav-novak", "hours (median)"),
        ("[SPACE] satisfaction_score", "Quarterly survey", 7.2, "2026-02-28", 90, "tomislav-novak", "/ 10"),
        ("[SPACE] performance_metric", "Custom domain metric", None, "2026-02-28", 90, "tomislav-novak", "domain-specific; populate per-firm"),
        ("[SPACE] activity_count", "GitHub + Slack", 124, "2026-02-28", 90, "tomislav-novak", "weekly average"),
        ("[SPACE] communication_quality", "Quarterly survey", 7.5, "2026-02-28", 90, "tomislav-novak", "/ 10"),
        ("[SPACE] efficiency_index", "Computed", 0.71, "2026-02-28", 90, "tomislav-novak", "0-1"),
        ("[DX Core 4] deep_work_hours", "Calendar analysis", 14.3, "2026-02-28", 90, "tomislav-novak", "weekly per partner"),
        ("[DX Core 4] cycle_time", "Linear/Jira", 5.2, "2026-02-28", 90, "tomislav-novak", "days (median)"),
        ("[DX Core 4] dx_score", "Quarterly survey", 6.8, "2026-02-28", 90, "tomislav-novak", "/ 10"),
        ("[DX Core 4] ai_assist_uptake", "Self-report weekly", 0.18, "2026-02-28", 90, "tomislav-novak", "fraction of commits AI-assisted"),
    ]
    for row in canonical_metrics:
        ws.append(row)
    _style_header_row(ws, last_col=7)
    _set_column_widths(ws, {"A": 36, "B": 22, "C": 14, "D": 14, "E": 18, "F": 18, "G": 40})

    # ── Sheet 2: Self_Report_vs_Actual ──
    ws2 = wb.create_sheet("Self_Report_vs_Actual")
    sra_headers = [
        "partner_id", "period", "self_reported_productivity",
        "measured_productivity", "gap", "gap_flag",
    ]
    for col_idx, h in enumerate(sra_headers, start=1):
        ws2.cell(row=1, column=col_idx, value=h)
    sample_rows = [
        ("[SAMPLE]davor-krznar", "week-4-rollout", -5, 5, None, None),
        ("[SAMPLE]mira-tek", "week-4-rollout", 30, 12, None, None),
        ("[SAMPLE]jana-kos", "week-4-rollout", 10, 8, None, None),
        ("[SAMPLE]mira-tek", "week-6-rollout", 35, 14, None, None),
    ]
    for row in sample_rows:
        ws2.append(row)
    for row_idx in range(2, 6):
        ws2.cell(row=row_idx, column=5, value=f"=C{row_idx}-D{row_idx}")
        ws2.cell(row=row_idx, column=6,
                 value=f'=IF(ABS(E{row_idx})>20,"PERCEPTION_GAP","OK")')
    ws2.conditional_formatting.add(
        "F2:F1000",
        CellIsRule(operator="equal", formula=['"PERCEPTION_GAP"'], fill=RED_FILL))
    # Comment in F1 about threshold provenance
    ws2["G1"] = "Threshold provenance"
    ws2["G2"] = ("20-point threshold ≈ half the 39-point swing observed in METR (2025) "
                 '"RCT on AI tools and developer productivity". Tighten to 15 once your firm has '
                 "90 days of internal baseline data.")
    _style_header_row(ws2, last_col=7)
    _set_column_widths(ws2, {"A": 22, "B": 22, "C": 24, "D": 22, "E": 12, "F": 18, "G": 80})

    # ── Sheet 3: Pilot_Cohort ──
    ws3 = wb.create_sheet("Pilot_Cohort")
    pc_headers = [
        "cohort_id", "partner_id", "inclusion_criteria", "start_date",
        "end_date", "role",
    ]
    for col_idx, h in enumerate(pc_headers, start=1):
        ws3.cell(row=1, column=col_idx, value=h)
    sample_cohort = [
        ("[SAMPLE]code-qa-2026-q1", "davor-krznar",
         "vibe-coder cohort; sceptical; 90-day baseline complete",
         "2026-03-01", "2026-04-26", "participant"),
        ("[SAMPLE]code-qa-2026-q1", "mira-tek",
         "agentic-engineer; enthusiastic; 90-day baseline complete",
         "2026-03-01", "2026-04-26", "participant"),
        ("[SAMPLE]code-qa-2026-q1", "jana-kos",
         "transitional cohort; curious; 90-day baseline complete",
         "2026-03-01", "2026-04-26", "participant"),
        ("[SAMPLE]code-qa-2026-q1", "tomislav-novak",
         "control; per-partner agreement; continues without Claude Code",
         "2026-03-01", "2026-04-26", "control"),
    ]
    for row in sample_cohort:
        ws3.append(row)
    dv_role = DataValidation(
        type="list",
        formula1='"participant,control,champion-candidate"', allow_blank=False)
    ws3.add_data_validation(dv_role)
    dv_role.add("F2:F1000")
    _style_header_row(ws3, last_col=6)

    # ── Sheet 4: Adoption_Curve ──
    ws4 = wb.create_sheet("Adoption_Curve")
    ac_headers = ["week_number", "active_users_pct",
                  "completion_rate", "satisfaction_score"]
    for col_idx, h in enumerate(ac_headers, start=1):
        ws4.cell(row=1, column=col_idx, value=h)
    for week, ap, cr, ss in [
        (1, 0.75, 0.60, 6.5), (2, 0.80, 0.70, 6.8), (3, 0.85, 0.78, 7.0),
        (4, 0.90, 0.82, 7.1), (5, 0.92, 0.85, 7.3), (6, 0.93, 0.88, 7.4),
        (7, 0.95, 0.90, 7.4), (8, 0.95, 0.92, 7.5),
    ]:
        ws4.append([week, ap, cr, ss])
    _style_header_row(ws4, last_col=4)

    # ── README ──
    readme = """# metr-baseline.xlsx

PURPOSE
Pre-rollout productivity baseline. MANDATORY before any major Skill Pack rollout per Skill Pack S6.

WRITTEN BY
- The pilot owner (typically the S6-running partner) at baseline capture (90 days pre-pilot).
- Weekly during the pilot (Self_Report_vs_Actual + Adoption_Curve).

READ BY
- Skill Pack S6 (Change Management).
- Routine R2 — surfaces perception-gap flags at Friday BR.

DESIGN DECISION — perception-gap threshold is 20 points
The threshold is roughly half the 39-point swing observed in METR (2025) "RCT on AI tools and developer productivity". Tighten to 15 once your firm has 90 days of internal baseline data.

DO NOT
- Roll out without a 90-day baseline. Mandatory per Skill Pack S6 §4.1 — no exceptions.
- Reconstruct a baseline from memory after-the-fact. The pack flags this as a discipline failure.
- Confront a perception gap directly. Surface the data; let the partner adjust (per S6 §4.6).
"""
    _add_metadata_sheet(wb, "X5")
    _add_readme_sheet(wb, readme)

    out_path = TEMPLATE_DIR / "metr-baseline.xlsx"
    wb.save(out_path)
    print(f"  ✓ {out_path.name}")


# ──────────────────────────────────────────────────────────────────────────────
# X6 — agent-skill-roi.xlsx
# ──────────────────────────────────────────────────────────────────────────────

def build_x6() -> None:
    """X6 agent-skill-roi.xlsx — agent costs vs. outputs."""
    wb = Workbook()
    wb.remove(wb.active)

    # ── Sheet 1: Agent_Costs ──
    ws = wb.create_sheet("Agent_Costs")
    ac_headers = [
        "date", "agent_or_skill", "input_tokens", "output_tokens",
        "api_calls", "cost_usd", "use_case",
    ]
    for col_idx, h in enumerate(ac_headers, start=1):
        ws.cell(row=1, column=col_idx, value=h)
    sample_costs = [
        ("2026-04-15", "claude-sonnet (compensation-attribution)", 12500, 3200, 8, 0.42, "R3 monthly variable"),
        ("2026-04-15", "claude-opus (klarna-test scoring)", 8200, 1800, 3, 0.81, "R7 KT-2026-014"),
        ("2026-04-15", "gemini-flash-lite (curator ingest)", 45000, 200, 12, 0.04, "S1 daily Brain ingest"),
    ]
    for row in sample_costs:
        ws.append(row)
    _style_header_row(ws, last_col=7)

    # ── Sheet 2: Skill_Outputs ──
    ws2 = wb.create_sheet("Skill_Outputs")
    so_headers = [
        "date", "skill_pack", "outputs_count",
        "hours_saved_estimated", "partner_value_uplift",
    ]
    for col_idx, h in enumerate(so_headers, start=1):
        ws2.cell(row=1, column=col_idx, value=h)
    ws2.append(["2026-04-15", "S3 compensation-attribution", 47, 12.5, 3100])
    ws2.append(["2026-04-15", "S4 code-qa", 23, 18.0, 8200])
    _style_header_row(ws2, last_col=5)

    # ── Sheet 3: Human_Agent_Ratio ──
    ws3 = wb.create_sheet("Human_Agent_Ratio")
    har_headers = [
        "domain", "human_partner_count", "ai_agent_count",
        "agent_human_ratio", "monthly_outputs_human",
        "monthly_outputs_agent_assisted", "ratio_trend",
    ]
    for col_idx, h in enumerate(har_headers, start=1):
        ws3.cell(row=1, column=col_idx, value=h)
    sample_har = [
        ("engineering", 8, 12, None, 64, 41, "+0.1 vs 30d ago"),
        ("sales", 3, 2, None, 14, 8, "0 vs 30d ago"),
        ("marketing", 2, 3, None, 22, 18, "+0.5 vs 30d ago"),
    ]
    for row_idx, row in enumerate(sample_har, start=2):
        for col_idx, val in enumerate(row, start=1):
            ws3.cell(row=row_idx, column=col_idx, value=val)
        # D = C / B
        ws3.cell(row=row_idx, column=4, value=f"=IFERROR(C{row_idx}/B{row_idx},0)")
    _style_header_row(ws3, last_col=7)

    # ── Sheet 4: ROI_Calc ──
    ws4 = wb.create_sheet("ROI_Calc")
    roi_headers = [
        "skill_pack", "monthly_cost", "monthly_outputs_value",
        "roi_multiple", "break_even_threshold",
    ]
    for col_idx, h in enumerate(roi_headers, start=1):
        ws4.cell(row=1, column=col_idx, value=h)
    ws4.append(["[SAMPLE]S3 compensation-attribution", 28, 4500, None, 1.0])
    ws4.append(["[SAMPLE]S4 code-qa", 142, 8200, None, 1.0])
    for row_idx in range(2, 4):
        ws4.cell(row=row_idx, column=4, value=f"=IFERROR(C{row_idx}/B{row_idx},0)")
    _style_header_row(ws4, last_col=5)

    # ── README ──
    readme = """# agent-skill-roi.xlsx

PURPOSE
Track per-Skill and per-agent costs vs. outputs. The "human-agent ratio" (Microsoft Frontier Firm metric) materialises in financial terms.

WRITTEN BY
- Routine R1 (daily output capture, indirectly via Skill_Outputs).
- Routine R8 (weekly Treasury Update — pulls Agent_Costs from API usage dashboards).
- Manual at Skill Pack rollouts (Skill_Outputs hours_saved_estimated).

READ BY
- Skill Pack S5 (Reporting & BR) — KPI movements block.
- Skill Pack S10 (Finance & Treasury) — financial close.
- Friday BR — KPI snapshot block.

DO NOT
- Change cost-attribution methodology without a Brain ADR (changes affect ROI calculations partners can see).
- Inflate `hours_saved_estimated` to make a Skill Pack rollout look favourable. Use the rolling 90-day baseline (per S6 §4.1) as the comparator.
"""
    _add_metadata_sheet(wb, "X6")
    _add_readme_sheet(wb, readme)

    out_path = TEMPLATE_DIR / "agent-skill-roi.xlsx"
    wb.save(out_path)
    print(f"  ✓ {out_path.name}")


# ──────────────────────────────────────────────────────────────────────────────
# X7 — eu-ai-act-mapping.xlsx
# ──────────────────────────────────────────────────────────────────────────────

def build_x7() -> None:
    """X7 eu-ai-act-mapping.xlsx — EU AI Act compliance register."""
    wb = Workbook()
    wb.remove(wb.active)

    # ── Sheet 1: Use_Cases ──
    ws = wb.create_sheet("Use_Cases")
    uc_headers = [
        "use_case_id", "name", "owner_partner_id", "brief_description",
        "deployment_status", "affected_population", "brain_link",
    ]
    for col_idx, h in enumerate(uc_headers, start=1):
        ws.cell(row=1, column=col_idx, value=h)
    sample_uc = [
        ("[SAMPLE]UC-001", "Code & QA AI assistance", "tomislav-novak",
         "AI-assisted code review and refactor across firm-saas repo",
         "production", "8 engineering partners", "[[architecture/code-qa-rollout-2026-q1]]"),
        ("[SAMPLE]UC-002", "Customer-onboarding email AI drafter", "davor-krznar",
         "AI drafts customer onboarding emails per template; opt-in escalation for high-tier",
         "production", "all customers (~87)", "[[klarna-tests/KT-2026-014]]"),
        ("[SAMPLE]UC-003", "Lumina widget RAG chatbot", "anya-gorska",
         "Customer-facing RAG chatbot front-door for inbound",
         "production", "all website visitors", "[[products/lumina-widget]]"),
    ]
    for row in sample_uc:
        ws.append(row)
    dv_status = DataValidation(
        type="list", formula1='"planned,pilot,production,retired"', allow_blank=False)
    ws.add_data_validation(dv_status)
    dv_status.add("E2:E1000")
    _style_header_row(ws, last_col=7)
    _set_column_widths(ws, {"A": 14, "B": 32, "C": 18, "D": 50, "E": 16, "F": 22, "G": 36})

    # ── Sheet 2: Annex_III_Risk_Mapping ──
    ws2 = wb.create_sheet("Annex_III_Risk_Mapping")
    arm_headers = [
        "use_case_id", "annex_iii_category", "rationale",
        "conservative_baseline_tier", "counsel_review_status",
        "counsel_review_date",
    ]
    for col_idx, h in enumerate(arm_headers, start=1):
        ws2.cell(row=1, column=col_idx, value=h)
    sample_arm = [
        ("[SAMPLE]UC-001", "none", "Internal engineering tooling; no decisions affecting persons.", "minimal", "approved", "2026-02-15"),
        ("[SAMPLE]UC-002", "Annex III §4 (employment/access)", "Touches customer-onboarding decisions; counsel ruled limited-risk on basis of human escalation path.", "limited", "approved", "2026-04-22"),
        ("[SAMPLE]UC-003", "Article 50 transparency", "Customer-facing chatbot; transparency obligation only.", "limited", "approved", "2026-01-30"),
    ]
    for row in sample_arm:
        ws2.append(row)
    dv_tier = DataValidation(type="list", formula1='"high,limited,minimal,prohibited"', allow_blank=False)
    ws2.add_data_validation(dv_tier)
    dv_tier.add("D2:D1000")
    dv_review = DataValidation(type="list", formula1='"pending,approved,objected,withdrawn"', allow_blank=False)
    ws2.add_data_validation(dv_review)
    dv_review.add("E2:E1000")
    _style_header_row(ws2, last_col=6)

    # ── Sheet 3: Article_Obligations ──
    ws3 = wb.create_sheet("Article_Obligations")
    ao_headers = [
        "use_case_id", "article_9_status", "article_12_status",
        "article_13_status", "article_14_status",
        "gdpr_article_22_status", "evidence_refs",
    ]
    for col_idx, h in enumerate(ao_headers, start=1):
        ws3.cell(row=1, column=col_idx, value=h)
    # only high-risk use cases need Article 9/12/13/14 mapping; sample data shows partial entries
    ws3.append([
        "[SAMPLE]UC-002 (limited-risk; partial)",
        "n/a (limited-risk)",
        "compliant (R6 audit trail)",
        "compliant (Skill Pack S3 limitations + pre-rollout consultation)",
        "compliant (founder + Mira non-beneficiary review)",
        "compliant (R3 requires human sign-off; no solely automated decision)",
        "[[klarna-tests/KT-2026-014]]; [[firm/audit-logs/2026-04-29]]",
    ])
    dv_article = DataValidation(
        type="list",
        formula1='"compliant,partial,not-started,n/a (limited-risk),n/a (minimal-risk)"', allow_blank=True)
    ws3.add_data_validation(dv_article)
    dv_article.add("B2:F1000")
    _style_header_row(ws3, last_col=7)

    # ── Sheet 4: Evidence_Trail ──
    ws4 = wb.create_sheet("Evidence_Trail")
    et_headers = [
        "use_case_id", "article", "requirement", "evidence_type",
        "evidence_link", "last_verified_date",
    ]
    for col_idx, h in enumerate(et_headers, start=1):
        ws4.cell(row=1, column=col_idx, value=h)
    ws4.append([
        "[SAMPLE]UC-002", "Article 12", "Daily logging of decisions",
        "routine_id", "R6 (daily 23:00)", "2026-04-29",
    ])
    dv_evtype = DataValidation(
        type="list",
        formula1='"skill_pack_ref,routine_id,brain_page,risk_register_row,external"', allow_blank=False)
    ws4.add_data_validation(dv_evtype)
    dv_evtype.add("D2:D1000")
    _style_header_row(ws4, last_col=6)

    # ── Sheet 5: Audit_Log_Index ──
    ws5 = wb.create_sheet("Audit_Log_Index")
    ali_headers = [
        "date", "audit_log_path", "entries_count", "anomalies_flagged",
    ]
    for col_idx, h in enumerate(ali_headers, start=1):
        ws5.cell(row=1, column=col_idx, value=h)
    ws5.append([
        "2026-04-29", "[[firm/audit-logs/2026-04-29]]", 14, "No",
    ])
    _style_header_row(ws5, last_col=4)

    # ── README ──
    readme = """# eu-ai-act-mapping.xlsx

PURPOSE
Compliance register for EU-operating organisations per governance/EU-AI-ACT.md. MANDATORY for EU adopters from 2 August 2026.

WRITTEN BY
- Skill Pack S7 (Governance & Compliance) at use-case classification time.
- Routine R6 (daily 23:00) — appends Audit_Log_Index row.
- Manual + counsel review for Annex_III_Risk_Mapping.

READ BY
- Skill Pack S7 at all times.
- Counsel — annual review minimum.
- Routine R2 (Weekly BR) — anomaly count surfaces in KPI block.

NOTE: this document is methodology, not legal advice. Counsel review is mandatory for EU adopters before relying on any classification herein.

DO NOT
- Treat compliance as a one-time exercise; it is continuous.
- Bypass the daily R6 audit trail.
- Claim EU AI Act compliance without counsel review.
- Modify a use case's Annex III classification without re-running counsel review.
"""
    _add_metadata_sheet(wb, "X7")
    _add_readme_sheet(wb, readme)

    out_path = TEMPLATE_DIR / "eu-ai-act-mapping.xlsx"
    wb.save(out_path)
    print(f"  ✓ {out_path.name}")


# ──────────────────────────────────────────────────────────────────────────────
# X8 — treasury-runway.xlsx (OPTIONAL)
# ──────────────────────────────────────────────────────────────────────────────

def build_x8() -> None:
    """X8 treasury-runway.xlsx — OPTIONAL; only orgs adopting Unit Fund."""
    wb = Workbook()
    wb.remove(wb.active)

    # ── Sheet 1: Cash_Position ──
    ws = wb.create_sheet("Cash_Position")
    cp_headers = ["date", "account_label", "balance", "currency", "balance_eur_equivalent"]
    for col_idx, h in enumerate(cp_headers, start=1):
        ws.cell(row=1, column=col_idx, value=h)
    ws.append(["2026-04-30", "Operating EUR", 240000, "EUR", 240000])
    ws.append(["2026-04-30", "Reserve EUR", 180000, "EUR", 180000])
    ws.append(["2026-04-30", "Stablecoin USDC", 25000, "USDC", 23500])
    _style_header_row(ws, last_col=5)

    # ── Sheet 2: Obligations ──
    ws2 = wb.create_sheet("Obligations")
    o_headers = ["due_date", "description", "amount", "currency", "recurring", "category"]
    for col_idx, h in enumerate(o_headers, start=1):
        ws2.cell(row=1, column=col_idx, value=h)
    ws2.append(["2026-05-05", "April variable + base payroll", 38000, "EUR", "yes", "payroll"])
    ws2.append(["2026-05-15", "Office rent (May)", 4200, "EUR", "yes", "suppliers"])
    ws2.append(["2026-07-01", "Q2 long-tail settlement (forecast)", 12000, "EUR", "yes", "long-tail"])
    dv_recur = DataValidation(type="list", formula1='"yes,no"', allow_blank=False)
    ws2.add_data_validation(dv_recur)
    dv_recur.add("E2:E1000")
    dv_cat = DataValidation(type="list",
                            formula1='"payroll,variable,long-tail,suppliers,other"', allow_blank=False)
    ws2.add_data_validation(dv_cat)
    dv_cat.add("F2:F1000")
    _style_header_row(ws2, last_col=6)

    # ── Sheet 3: Runway_Calc ──
    ws3 = wb.create_sheet("Runway_Calc")
    rc_headers = [
        "snapshot_date", "total_cash_eur", "monthly_burn_average",
        "runway_months", "unit_fund_outstanding_units",
        "implied_redemption_value", "reserve_coverage_ratio",
    ]
    for col_idx, h in enumerate(rc_headers, start=1):
        ws3.cell(row=1, column=col_idx, value=h)
    # B (total_cash_eur) = SUM Cash_Position E
    # C (monthly_burn_average) — comment: derived from historical Cash_Position deltas, not Obligations.
    # D (runway_months) = B / C
    ws3.append([
        "2026-04-30", 443500, 24500, None,
        0, 0, None,
    ])
    ws3["D2"] = "=IFERROR(B2/C2,0)"
    ws3["G2"] = '=IFERROR(B2/IF(F2=0,1,F2),0)'
    # Conditional formatting: D red if <6 months — guard on populated row (A non-blank)
    # so blank rows (CellIs treats empty as 0) are not painted red.
    ws3.conditional_formatting.add(
        "D2:D1000",
        FormulaRule(formula=["AND($A2<>\"\",D2<6)"], fill=RED_FILL))
    # G red if <1.0 — same populated-row guard.
    ws3.conditional_formatting.add(
        "G2:G1000",
        FormulaRule(formula=["AND($A2<>\"\",G2<1)"], fill=RED_FILL))
    # Comment / note about burn-average source
    ws3["H1"] = "Note"
    ws3["H2"] = ("monthly_burn_average is derived from rolling 3-month Cash_Position deltas "
                 "(realised outflow), NOT from the future-committed total in Obligations. "
                 "R8 writes this on each run.")
    _style_header_row(ws3, last_col=8)

    # ── Sheet 4: Reserve_Discipline ──
    ws4 = wb.create_sheet("Reserve_Discipline")
    rd_headers = ["parameter", "value", "last_review_date", "owner"]
    for col_idx, h in enumerate(rd_headers, start=1):
        ws4.cell(row=1, column=col_idx, value=h)
    ws4.append(["minimum_reserve_months", 9, "2026-01-15", "founder-tali"])
    ws4.append(["unit_fund_redemption_pct_per_quarter_max", 0.05, "2026-01-15", "founder-tali"])
    ws4.append(["bid_update_cadence", "monthly", "2026-01-15", "founder-tali"])
    _style_header_row(ws4, last_col=4)

    # ── README ──
    readme = """# treasury-runway.xlsx (OPTIONAL)

PURPOSE
Treasury reserve discipline. OPTIONAL in v1.0 — required only for organisations adopting the Unit Fund (Generation 2). Absent the Unit Fund, simpler treasury tracking via accountant + bank statements is sufficient.

DESIGN DECISION — monthly_burn_average is from historical Cash_Position deltas
NOT from future-committed Obligations. Burn is realised outflow over the rolling 3 months. R8 writes this on each run. Obligations forecasts forward; Cash_Position records what actually happened.

WRITTEN BY
- Routine R8 (Treasury Runway Update, weekly Monday 08:00) — populates Cash_Position + Obligations + Runway_Calc.
- Manual at Reserve_Discipline parameter changes (founder authority).

READ BY
- Skill Pack S10 (Finance & Treasury).
- Skill Pack S5 (Reporting & BR) — runway_months in KPI block.

DO NOT
- Open the Unit Fund without ≥6-9 months of attribution-accuracy data + reserve_coverage_ratio ≥1.0.
- Change Reserve_Discipline parameters without a Brain ADR.
- Bypass treasury runway discipline once the Unit Fund is open.
"""
    _add_metadata_sheet(wb, "X8")
    _add_readme_sheet(wb, readme)

    out_path = TEMPLATE_DIR / "treasury-runway.xlsx"
    wb.save(out_path)
    print(f"  ✓ {out_path.name}")


# ──────────────────────────────────────────────────────────────────────────────
# X9 — oot-readiness.xlsx
# ──────────────────────────────────────────────────────────────────────────────

def build_x9() -> None:
    """X9 oot-readiness.xlsx — 20-question pre-adoption diagnostic."""
    wb = Workbook()
    wb.remove(wb.active)

    # ── Sheet 1: Questions ──
    ws = wb.create_sheet("Questions")
    q_headers = [
        "question_id", "dimension", "question",
        "scoring_guidance", "response", "evidence",
    ]
    for col_idx, h in enumerate(q_headers, start=1):
        ws.cell(row=1, column=col_idx, value=h)

    questions: list[tuple[str, str, str, str]] = [
        # People (Q01–Q05)
        ("Q01", "People",
         "Are 80%+ of intended partners willing to operate without fixed hours?",
         "0=No / 1=Mixed / 2=Yes (written buy-in from each)"),
        ("Q02", "People",
         "Is there written buy-in for output-based variable compensation?",
         "0=No / 1=Verbal only / 2=Written sign-off"),
        ("Q03", "People",
         "Are reward-species expectations aligned across the founding partners?",
         "0=Not discussed / 1=Discussed; misaligned / 2=Discussed; aligned with renegotiation cadence agreed"),
        ("Q04", "People",
         "Is the founder willing to stop tracking time and headcount as primary metrics?",
         "0=No / 1=Verbally yes; not committed / 2=Committed in writing"),
        ("Q05", "People",
         "Have all founding partners read governance/DECISION-RIGHTS.md and accepted the three-tier dispute path?",
         "0=Not read / 1=Read; not all accepted / 2=Read and accepted by all"),
        # Tech (Q06–Q10)
        ("Q06", "Tech",
         "Is Claude Pro/Max budget approved (cloud track) OR is there an always-on machine plan (privacy track)?",
         "0=No / 1=Partial / 2=Full"),
        ("Q07", "Tech",
         "Is Bitwarden + Trezor procurement approved?",
         "0=No / 1=Bitwarden only / 2=Both"),
        ("Q08", "Tech",
         "Is the Curator install path clear for at least one partner?",
         "0=No / 1=Clear; not run / 2=Installed and self-test passed"),
        ("Q09", "Tech",
         "Are GitHub + Drive (cloud) or 4thtech + PollinationX (privacy) workflows acceptable to all partners?",
         "0=Resistance from some / 1=Mixed / 2=Acceptable to all"),
        ("Q10", "Tech",
         "Has the Skill Pack ecosystem been demoed to all founding partners?",
         "0=No / 1=Some / 2=All"),
        # Culture (Q11–Q15)
        ("Q11", "Culture",
         "Is there willingness to run a METR baseline before any Skill rollout?",
         "0=No / 1=Reluctant / 2=Yes"),
        ("Q12", "Culture",
         "Is the Klarna Test framing accepted as the framework's signature discipline?",
         "0=Not understood / 1=Understood; not committed / 2=Committed"),
        ("Q13", "Culture",
         "Are the founding partners comfortable with public commitment to long-tail entitlements?",
         "0=No / 1=Some / 2=All"),
        ("Q14", "Culture",
         "Is the Friday Business Review structure compatible with existing rhythms?",
         "0=No / 1=Adjustable / 2=Yes"),
        ("Q15", "Culture",
         "Is the founder willing to defend ØØT-native decisions publicly in 2 years?",
         "0=No / 1=Maybe / 2=Yes"),
        # Risk (Q16–Q20)
        ("Q16", "Risk",
         "Has local counsel been identified and engaged for the eleven legal touchpoints?",
         "0=No / 1=Identified; not engaged / 2=Engaged"),
        ("Q17", "Risk",
         "Is the jurisdiction's worker classification compatible with partner-not-employee?",
         "0=No / 1=Counsel needs to verify / 2=Counsel verified"),
        ("Q18", "Risk",
         "Are EU AI Act obligations understood (if EU-operating)?",
         "0=No or n/a (not EU) / 1=Understood at high level / 2=Mapped per use case"),
        ("Q19", "Risk",
         "Is the secrets policy (Bitwarden + Trezor) implementable culturally?",
         "0=No / 1=Some resistance / 2=Yes"),
        ("Q20", "Risk",
         "Is there budget for Curator pay-as-you-go (~€10/month/heavy-user) and Bitwarden/Trezor procurement?",
         "0=No / 1=Partial / 2=Yes"),
    ]
    for q_id, dim, q, guidance in questions:
        ws.append([q_id, dim, q, guidance, None, None])
    # Sample responses (founder fills in)
    sample_scores = [2, 1, 1, 2, 2,  # People
                     2, 2, 2, 2, 1,  # Tech
                     2, 2, 1, 2, 2,  # Culture
                     1, 2, 2, 2, 2]  # Risk
    for i, score in enumerate(sample_scores, start=2):
        ws.cell(row=i, column=5, value=score)

    dv_score = DataValidation(type="list", formula1='"0,1,2"', allow_blank=True)
    ws.add_data_validation(dv_score)
    dv_score.add("E2:E21")

    _style_header_row(ws, last_col=6)
    _set_column_widths(ws, {"A": 8, "B": 12, "C": 60, "D": 60, "E": 10, "F": 30})

    # ── Sheet 2: Scoring ──
    ws2 = wb.create_sheet("Scoring")
    s_headers = ["dimension", "dimension_score", "dimension_max", "dimension_pct"]
    for col_idx, h in enumerate(s_headers, start=1):
        ws2.cell(row=1, column=col_idx, value=h)
    dimensions = [
        ("People", "Questions!E2:E6"),
        ("Tech", "Questions!E7:E11"),
        ("Culture", "Questions!E12:E16"),
        ("Risk", "Questions!E17:E21"),
    ]
    for row_idx, (dim, rng) in enumerate(dimensions, start=2):
        ws2.cell(row=row_idx, column=1, value=dim)
        ws2.cell(row=row_idx, column=2, value=f"=SUM({rng})")
        ws2.cell(row=row_idx, column=3, value=10)
        ws2.cell(row=row_idx, column=4, value=f"=B{row_idx}/C{row_idx}")
    # Total row
    ws2.append(["TOTAL", None, 40, None])
    ws2["B6"] = "=SUM(B2:B5)"
    ws2["D6"] = "=B6/C6"
    # Conditional formatting: D6 (total_pct) — green ≥75%, yellow 60–<75%, red <60%.
    # Green is added FIRST so it takes priority at exactly 0.75 (the >=0.75 and the
    # 0.60–0.75 ranges both matched at the boundary; earlier-added rule wins in Excel).
    ws2.conditional_formatting.add(
        "D6", CellIsRule(operator="greaterThanOrEqual", formula=["0.75"], fill=GREEN_FILL))
    ws2.conditional_formatting.add(
        "D6", CellIsRule(operator="between", formula=["0.60", "0.7499"], fill=YELLOW_FILL))
    ws2.conditional_formatting.add(
        "D6", CellIsRule(operator="lessThan", formula=["0.60"], fill=RED_FILL))
    # Format D column as percentage
    for row_idx in range(2, 7):
        ws2.cell(row=row_idx, column=4).number_format = "0%"
    _style_header_row(ws2, last_col=4)

    # ── Sheet 3: Recommendations ──
    ws3 = wb.create_sheet("Recommendations")
    r_headers = ["dimension", "recommendation"]
    for col_idx, h in enumerate(r_headers, start=1):
        ws3.cell(row=1, column=col_idx, value=h)
    recs = [
        ("People", "If <60%: defer adoption until written partnership-buy-in exists and reward-species alignment is documented. See Skill Pack S3 §4.1."),
        ("Tech", "If <60%: complete the Curator install + Bitwarden + Trezor procurement before adoption. See QUICKSTART.md and governance/SECRETS-POLICY.md."),
        ("Culture", "If <60%: run a 1-week Klarna Test framing workshop. Read MANIFESTO.md Thesis 1 together. Cultural readiness is the bottleneck, not technical."),
        ("Risk", "If <60%: engage local counsel before any partner contract is signed. See docs/06-when-to-call-a-lawyer.md."),
    ]
    for row in recs:
        ws3.append(row)
    _style_header_row(ws3, last_col=2)
    _set_column_widths(ws3, {"A": 14, "B": 100})

    # ── README ──
    readme = """# oot-readiness.xlsx

PURPOSE
20-question pre-adoption diagnostic across People / Tech / Culture / Risk dimensions. Run once before adopting ØØT.

INTERPRETATION
- Total ≥75% (≥30/40): green light. Adopt.
- Total 60–75% (24-29/40): yellow. Address gaps in lowest-scoring dimension first.
- Total <60% (<24/40): red. Adoption is premature. The framework's discipline applies to the framework itself: don't push past genuine readiness gaps.

WRITTEN BY
- The founder, once at adoption.
- Re-run quarterly during Year 1 to track readiness evolution; annually thereafter.

REVIEWED BY
- Founder.
- Recommended: read the recommendations sheet alongside the score; address weakest dimension first.

DO NOT
- Score yourself optimistically. The discipline of the framework starts with the readiness assessment.
- Skip questions. Each question maps to a real adoption risk.
- Adopt below 60% without written documentation of how the gaps will be addressed in the first 90 days.

This is a diagnostic, not a gate. Founders may proceed below 60% with documented mitigation, but the framework's authors recommend honestly addressing gaps before sunk-cost dynamics make later course correction harder.
"""
    _add_metadata_sheet(wb, "X9")
    _add_readme_sheet(wb, readme)

    out_path = TEMPLATE_DIR / "oot-readiness.xlsx"
    wb.save(out_path)
    print(f"  ✓ {out_path.name}")


# ──────────────────────────────────────────────────────────────────────────────
# Main
# ──────────────────────────────────────────────────────────────────────────────

def main() -> None:
    """Generate all 9 Excel templates."""
    print(f"Generating ØØT Excel templates → {TEMPLATE_DIR}")
    TEMPLATE_DIR.mkdir(parents=True, exist_ok=True)

    builders = [
        ("X1 partner-output-ledger", build_x1),
        ("X2 reward-species-declaration", build_x2),
        ("X3 business-review", build_x3),
        ("X4 klarna-test", build_x4),
        ("X5 metr-baseline", build_x5),
        ("X6 agent-skill-roi", build_x6),
        ("X7 eu-ai-act-mapping", build_x7),
        ("X8 treasury-runway", build_x8),
        ("X9 oot-readiness", build_x9),
    ]
    for name, fn in builders:
        print(f"\n{name}:")
        fn()

    print(f"\n✓ All 9 templates generated in {TEMPLATE_DIR}")


if __name__ == "__main__":
    main()

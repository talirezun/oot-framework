"""Structural tests for the committed Excel templates + build_excel.py.

These assert against the COMMITTED templates/excel/*.xlsx files. They do NOT run
the generator by default (that would mutate the working tree). One opt-in test
regenerates and restores via git checkout in a finally block.

Contracts checked (see templates/excel/SPEC.md):
  - All 9 workbooks exist and open.
  - X1 (partner-output-ledger) Output_Log has K2/L2 formulas
    (value_envelope VLOOKUP + computed_variable with the rework zero-out).
  - X4 (klarna-test) Klarna_Score L2 == "=SUM(B2:K2)" and the >=14 threshold
    is present (Decision_Log I2 formula + Klarna_Score conditional formatting).
  - X9 (oot-readiness) Scoring dimension SUM formulas exist.
"""

from __future__ import annotations

import subprocess
from pathlib import Path

import pytest
from openpyxl import load_workbook

REPO_ROOT = Path(__file__).resolve().parent.parent
EXCEL_DIR = REPO_ROOT / "templates" / "excel"

EXPECTED_WORKBOOKS = [
    "partner-output-ledger.xlsx",
    "reward-species-declaration.xlsx",
    "business-review.xlsx",
    "klarna-test.xlsx",
    "metr-baseline.xlsx",
    "agent-skill-roi.xlsx",
    "eu-ai-act-mapping.xlsx",
    "treasury-runway.xlsx",
    "oot-readiness.xlsx",
]


def test_build_excel_importable():
    """The generator module must import cleanly (used by CI + oot-build-excel)."""
    import importlib.util

    path = REPO_ROOT / "scripts" / "build_excel.py"
    spec = importlib.util.spec_from_file_location("build_excel", path)
    module = importlib.util.module_from_spec(spec)
    assert spec and spec.loader
    spec.loader.exec_module(module)
    assert hasattr(module, "main")
    # It knows where to write.
    assert module.TEMPLATE_DIR.name == "excel"


@pytest.mark.parametrize("name", EXPECTED_WORKBOOKS)
def test_workbook_exists_and_opens(name):
    path = EXCEL_DIR / name
    assert path.exists(), f"{name} missing from templates/excel/"
    wb = load_workbook(path, data_only=False)
    assert "_metadata" in wb.sheetnames, f"{name} missing hidden _metadata sheet"
    assert "README" in wb.sheetnames, f"{name} missing README sheet"


def test_all_nine_present():
    found = sorted(p.name for p in EXCEL_DIR.glob("*.xlsx"))
    for name in EXPECTED_WORKBOOKS:
        assert name in found, f"{name} not found in {found}"
    assert len(found) == 9, f"expected exactly 9 workbooks, found {len(found)}: {found}"


def test_x1_output_log_formulas():
    """X1 Output_Log K2 = value_envelope VLOOKUP; L2 = computed_variable with rework zero-out + weight."""
    wb = load_workbook(EXCEL_DIR / "partner-output-ledger.xlsx", data_only=False)
    ws = wb["Output_Log"]

    k2 = str(ws["K2"].value or "")
    assert k2.startswith("=VLOOKUP("), f"X1 K2 not a VLOOKUP: {k2!r}"
    assert "$O$2:$P$5" in k2, f"X1 K2 does not reference the envelope table: {k2!r}"

    l2 = str(ws["L2"].value or "")
    assert l2.startswith("="), f"X1 L2 not a formula: {l2!r}"
    assert "K2" in l2 and "J2" in l2, f"X1 L2 missing K/J references: {l2!r}"
    # ADR-005: the co-authorship weight (column N) must factor into computed_variable.
    assert "N2" in l2, f"X1 L2 missing weight (N2) factor: {l2!r}"
    # The rework-within-30d zero-out (I2="Yes" => 0).
    assert 'IF(I2="Yes",0,1)' in l2.replace(" ", ""), f"X1 L2 missing rework zero-out: {l2!r}"


def test_x1_weight_column():
    """X1 Output_Log gains column N `weight`, default 1.0 on sample rows (ADR-005)."""
    wb = load_workbook(EXCEL_DIR / "partner-output-ledger.xlsx", data_only=False)
    ws = wb["Output_Log"]
    assert ws["N1"].value == "weight", f"X1 N1 not 'weight': {ws['N1'].value!r}"
    assert float(ws["N2"].value) == 1.0, f"X1 N2 default weight not 1.0: {ws['N2'].value!r}"


def test_x4_status_column():
    """X4 Decision_Log gains column M `status` with the lifecycle enum validation (ADR-004)."""
    wb = load_workbook(EXCEL_DIR / "klarna-test.xlsx", data_only=False)
    dl = wb["Decision_Log"]
    assert dl["M1"].value == "status", f"X4 Decision_Log M1 not 'status': {dl['M1'].value!r}"
    # status is a literal, never a formula.
    m2 = str(dl["M2"].value or "")
    assert m2 and not m2.startswith("="), f"X4 M2 should be a literal enum value, got: {m2!r}"

    # A data-validation list covering column M enumerates the five lifecycle states.
    m_lists = [
        str(dv.formula1)
        for dv in dl.data_validations.dataValidation
        if any("M" in str(rng) for rng in dv.sqref.ranges)
    ]
    assert any(
        all(state in lst for state in ("scoring", "remediation", "monitoring", "proceeded", "held"))
        for lst in m_lists
    ), f"X4 Decision_Log status validation missing lifecycle enum: {m_lists}"


def test_x2_partner_id_join_key():
    """X2 Base_Variable_Split and Long_Tail_Schedule gain a leading partner_id column (ADR-005)."""
    wb = load_workbook(EXCEL_DIR / "reward-species-declaration.xlsx", data_only=False)
    for sheet in ("Base_Variable_Split", "Long_Tail_Schedule"):
        ws = wb[sheet]
        assert ws["A1"].value == "partner_id", f"X2 {sheet} A1 not 'partner_id': {ws['A1'].value!r}"
        # Sample single-partner firm backfills P-001 (ADR-005 §3).
        assert (
            ws["A2"].value == "P-001"
        ), f"X2 {sheet} A2 sample partner_id not 'P-001': {ws['A2'].value!r}"
    # The variable weights are now in D/E/F (shifted +1 by the leading column) and sum to 1.0.
    bvs = wb["Base_Variable_Split"]
    weights = [bvs["D2"].value, bvs["E2"].value, bvs["F2"].value]
    assert round(sum(weights), 3) == 1.0, f"X2 shifted weights D/E/F don't sum to 1.0: {weights}"


def test_x4_klarna_score_total_and_threshold():
    """X4 Klarna_Score L2 == '=SUM(B2:K2)'; the >=14 threshold is present."""
    wb = load_workbook(EXCEL_DIR / "klarna-test.xlsx", data_only=False)
    ks = wb["Klarna_Score"]

    l2 = str(ks["L2"].value or "").replace(" ", "")
    assert l2 == "=SUM(B2:K2)", f"X4 Klarna_Score L2 not =SUM(B2:K2): {l2!r}"

    # data_only must be None (openpyxl writes no cached value) — this is exactly
    # why the klarna-gate workflow needs a recompute fallback.
    wb_cached = load_workbook(EXCEL_DIR / "klarna-test.xlsx", data_only=True)
    assert wb_cached["Klarna_Score"]["L2"].value is None

    # Threshold 14 lives in Decision_Log I2 (PROCEED/HOLD) ...
    dl = wb["Decision_Log"]
    i2 = str(dl["I2"].value or "")
    assert "H2>=14" in i2.replace(" ", ""), f"X4 Decision_Log I2 missing >=14 threshold: {i2!r}"

    # ... and in the Klarna_Score total conditional formatting.
    cf_formulas = []
    for rng, rules in ks.conditional_formatting._cf_rules.items():
        if "L2" in str(rng.sqref):
            for rule in rules:
                cf_formulas.extend(str(f) for f in (rule.formula or []))
    assert any(
        "14" in f for f in cf_formulas
    ), f"X4 Klarna_Score total conditional formatting does not reference 14: {cf_formulas}"


def test_x9_dimension_sums():
    """X9 Scoring sheet has per-dimension SUM formulas + a grand total."""
    wb = load_workbook(EXCEL_DIR / "oot-readiness.xlsx", data_only=False)
    sc = wb["Scoring"]

    # B2..B5 = per-dimension sums over the Questions responses.
    for r in range(2, 6):
        val = str(sc.cell(row=r, column=2).value or "")
        assert val.startswith("=SUM(Questions!"), f"X9 Scoring B{r} not a dimension SUM: {val!r}"
    # B6 = grand total across the four dimensions.
    b6 = str(sc["B6"].value or "").replace(" ", "")
    assert b6 == "=SUM(B2:B5)", f"X9 Scoring B6 not the grand total: {b6!r}"


def test_regenerate_matches_committed_optional():
    """Opt-in: regenerate and confirm no drift, then restore the committed tree.

    Skipped unless OOT_RUN_GENERATOR=1 so the default suite never mutates the
    working tree. Mirrors the excel-validation.yml drift check.
    """
    import os

    if os.environ.get("OOT_RUN_GENERATOR") != "1":
        pytest.skip("set OOT_RUN_GENERATOR=1 to run the mutating drift check")

    import shutil
    import tempfile
    import importlib.util

    backup = Path(tempfile.mkdtemp(prefix="oot-committed-xlsx-"))
    try:
        for p in EXCEL_DIR.glob("*.xlsx"):
            shutil.copy2(p, backup / p.name)

        spec = importlib.util.spec_from_file_location(
            "build_excel", REPO_ROOT / "scripts" / "build_excel.py"
        )
        module = importlib.util.module_from_spec(spec)
        spec.loader.exec_module(module)
        module.main()

        diffs = []
        for name in EXPECTED_WORKBOOKS:
            a = load_workbook(backup / name, data_only=False)
            b = load_workbook(EXCEL_DIR / name, data_only=False)
            assert a.sheetnames == b.sheetnames, f"{name}: sheet names drifted"
            for sheet in a.sheetnames:
                wsa, wsb = a[sheet], b[sheet]
                mr = max(wsa.max_row, wsb.max_row)
                mc = max(wsa.max_column, wsb.max_column)
                for r in range(1, mr + 1):
                    for c in range(1, mc + 1):
                        if sheet == "_metadata" and r == 3 and c == 2:
                            continue  # generation_date legitimately changes
                        va = wsa.cell(row=r, column=c).value
                        vb = wsb.cell(row=r, column=c).value
                        if va != vb:
                            coord = wsa.cell(row=r, column=c).coordinate
                            diffs.append(f"{name}[{sheet}]!{coord}: {va!r} != {vb!r}")
        assert not diffs, "drift detected:\n" + "\n".join(diffs[:50])
    finally:
        # Restore the committed workbooks so the working tree is left clean.
        # Scope the restore to *.xlsx ONLY — a bare `templates/excel/` would
        # also revert uncommitted edits to templates/excel/SPEC.md.
        subprocess.run(
            ["git", "checkout", "--", "templates/excel/*.xlsx"],
            cwd=REPO_ROOT,
            check=False,
        )
        shutil.rmtree(backup, ignore_errors=True)
